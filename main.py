from icalendar import Calendar, Event
import openpyxl
import datetime
import re
import os
import csv

INPUT_FILE = "plan.xlsx"
OUTPUT_FILE_CSV_LOAD = "loaded_data.csv"
OUTPUT_FILE_CSV_CALENDAR = "csv_scheme.csv"
OUTPUT_FILE_ICS_CALENDAR = "calendar_scheme.ics"
OUTPUT_FILE_XLSX = "extracted_plan.xlsx"

HOURS = {
    "1": {"start": "8:00", "end": "9:30"},
    "3": {"start": "9:40", "end": "11:10"},
    "5": {"start": "11:30", "end": "13:00"},
    "7": {"start": "13:10", "end": "14:40"},
    "9": {"start": "14:45", "end": "16:15"},
    "11": {"start": "16:20", "end": "17:50"},
    "13": {"start": "17:55", "end": "19:25"},
    "15": {"start": "19:30", "end": "21:00"},
}

REPLACE_MAP = {
    "Ą": "A",
    "Ć": "C",
    "Ę": "E",
    "Ł": "L",
    "Ń": "N",
    "Ó": "O",
    "Ś": "S",
    "Ż": "Z",
    "Ź": "Z",
}

MAJORS = {
    "inf": "- informatyka",
    "log": "i stopnia - logistyka",
    "bezp": "i stopnia - kierunek bezp",
}

WEB_TO_MAJORS = {
    "informatyka": {
        "prefix": "inf",
        "searchParam": MAJORS["inf"],
    },
    "logistyka": {
        "prefix": "log",
        "searchParam": MAJORS["log"],
    },
    "bezpieczenstwo": {
        "prefix": "bezp",
        "searchParam": MAJORS["bezp"],
    },
}


def retrieve_start(hour):
    return HOURS[str(hour)]["start"]


def retrieve_end(hour):
    return HOURS[str(hour)]["end"]


def replace_polish_chars(string):
    for char, changed_char in REPLACE_MAP.items():
        string = string.replace(char, changed_char)
    return string


def get_calendar_schemes(wb):
    for prefix, major in MAJORS.items():
        majorOffsets = get_offsets(wb, major)
        data_loader(wb, majorOffsets, prefix)


def get_major_calendar_scheme(wb, major, year):
    majorOffsets, prefix = get_major_offset(wb, major)
    data_loader_for_specific_major(wb, majorOffsets, prefix, year)


def csv_to_ics(csv_file, prefix, yearId):
    calendar_folder = "calendars"
    if not os.path.exists(calendar_folder):
        os.makedirs(calendar_folder)

    ics_file = os.path.join(
        calendar_folder, f"{prefix}_{yearId}_{OUTPUT_FILE_ICS_CALENDAR}"
    )
    print(f"Creating calendar for {ics_file}...")
    with open(csv_file, "r") as file:
        csv_reader = csv.DictReader(file)

        cal = Calendar()
        for row in csv_reader:
            event = Event()

            event.add("summary", row["Subject"])
            event.add("description", row["Description"])
            event.add("location", row["Location"])

            start_datetime = datetime.datetime.strptime(
                row["Start Date"] + " " + row["Start Time"], "%Y-%m-%d %H:%M"
            )
            end_datetime = datetime.datetime.strptime(
                row["End Date"] + " " + row["End Time"], "%Y-%m-%d %H:%M"
            )
            event.add("dtstart", start_datetime)
            event.add("dtend", end_datetime)

            cal.add_component(event)

    with open(ics_file, "wb") as file:
        file.write(cal.to_ical())
    os.remove(csv_file)


def create_workbook(input_file, save_file=False, output_file=OUTPUT_FILE_XLSX):
    print("Creating workbook from excel...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active

    for row in ws.iter_rows(values_only=True):
        ws_new.append(row)

    if save_file:
        print("Excel saved to ", output_file)
        wb_new.save(output_file)

    return wb_new


def get_offsets(workbook, major):
    print("Finding offsets...")
    ret = {
        "DATA": {
            "ROW": 0,
            "COL_1": 0,
            "COL_2": 0,
        },
        "DATA_OFFSET": 0,
        "CLASS_START_COL": 0,
        "CLASS_END_COl": 15,
        "DATA_CLASS_OFFSET": 0,
        "CLASSES_SECTION": {},
    }
    ws = workbook.active
    row_number = 1
    for row in ws.iter_rows(values_only=True):
        cell_number = 1
        for cell_value in row:
            cell_val = str(cell_value).lower().strip().replace("  ", " ")
            next_cell_val = (
                str(ws.cell(row=row_number, column=cell_number + 1).value)
                .lower()
                .strip()
            )
            # class col
            if ret["CLASS_START_COL"] == 0 and cell_val == "1" and next_cell_val == "2":
                ret["CLASS_START_COL"] = cell_number

            # class section check
            if major in cell_val:
                year = cell_value.split(" ")[0]
                if year.isdigit():
                    if not ret["CLASSES_SECTION"].get(year):
                        ret["CLASSES_SECTION"][year] = [row_number + 1, row_number + 7]

            # check for data row
            try:
                if isinstance(
                    datetime.datetime.strptime(cell_val, "%Y-%m-%d %H:%M:%S"),
                    datetime.datetime,
                ):
                    if ret["DATA"]["ROW"] == 0:
                        ret["DATA"]["ROW"] = row_number
                    if ret["DATA"]["COL_1"] == 0:
                        ret["DATA"]["COL_1"] = cell_number
                    elif ret["DATA"]["COL_2"] == 0:
                        ret["DATA"]["COL_2"] = cell_number
            except ValueError:
                pass
            cell_number += 1
        row_number += 1
    if ret["DATA"]["COL_1"] != 0 and ret["DATA"]["COL_2"] != 0:
        ret["DATA_OFFSET"] = abs(ret["DATA"]["COL_1"] - ret["DATA"]["COL_2"])

    if ret["CLASS_START_COL"] != 0:
        ret["DATA_CLASS_OFFSET"] = abs(ret["CLASS_START_COL"] - ret["DATA"]["COL_1"])
    #  print(ret)
    return ret


def get_major_offset(workbook, major):
    searchParam = WEB_TO_MAJORS[major.lower()]["searchParam"]
    prefix = WEB_TO_MAJORS[major.lower()]["prefix"]
    print(f"Finding offset for {major}...")
    ret = {
        "DATA": {
            "ROW": 0,
            "COL_1": 0,
            "COL_2": 0,
        },
        "DATA_OFFSET": 0,
        "CLASS_START_COL": 0,
        "CLASS_END_COl": 15,
        "DATA_CLASS_OFFSET": 0,
        "CLASSES_SECTION": {},
    }
    ws = workbook.active
    row_number = 1
    for row in ws.iter_rows(values_only=True):
        cell_number = 1
        for cell_value in row:
            cell_val = str(cell_value).lower().strip().replace("  ", " ")
            next_cell_val = (
                str(ws.cell(row=row_number, column=cell_number + 1).value)
                .lower()
                .strip()
            )
            # class col
            if ret["CLASS_START_COL"] == 0 and cell_val == "1" and next_cell_val == "2":
                ret["CLASS_START_COL"] = cell_number

            # class section check
            if searchParam in cell_val:
                year = cell_value.split(" ")[0]
                if year.isdigit():
                    if not ret["CLASSES_SECTION"].get(year):
                        ret["CLASSES_SECTION"][year] = [row_number + 1, row_number + 7]

            # check for data row
            try:
                if isinstance(
                    datetime.datetime.strptime(cell_val, "%Y-%m-%d %H:%M:%S"),
                    datetime.datetime,
                ):
                    if ret["DATA"]["ROW"] == 0:
                        ret["DATA"]["ROW"] = row_number
                    if ret["DATA"]["COL_1"] == 0:
                        ret["DATA"]["COL_1"] = cell_number
                    elif ret["DATA"]["COL_2"] == 0:
                        ret["DATA"]["COL_2"] = cell_number
            except ValueError:
                pass
            cell_number += 1
        row_number += 1
    if ret["DATA"]["COL_1"] != 0 and ret["DATA"]["COL_2"] != 0:
        ret["DATA_OFFSET"] = abs(ret["DATA"]["COL_1"] - ret["DATA"]["COL_2"])

    if ret["CLASS_START_COL"] != 0:
        ret["DATA_CLASS_OFFSET"] = abs(ret["CLASS_START_COL"] - ret["DATA"]["COL_1"])
    return ret, prefix


def data_loader_for_specific_major(workbook, offsets, prefix, year):
    sheet = workbook.active

    print(f"Loading data for year {prefix}_{year}...")
    output_file = OUTPUT_FILE_CSV_LOAD + f"_{year}"
    matrix = []
    for start_col in range(
        offsets["CLASS_START_COL"], sheet.max_column, offsets["DATA_OFFSET"]
    ):
        value = sheet.cell(
            offsets["DATA"]["ROW"], column=start_col + offsets["DATA_CLASS_OFFSET"]
        ).value
        data = str(value).split(" ")[0]
        matrix.append([data])

        end_col = start_col + offsets["CLASS_END_COl"]
        for row in range(
            offsets["CLASSES_SECTION"][str(year)][0] + 1,
            offsets["CLASSES_SECTION"][str(year)][1] + 1,
        ):
            row_data = []
            for col in range(start_col, min(end_col + 1, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col).value
                data = str(cell)
                row_data.append(data)
            matrix.append(row_data)

    with open(output_file, "w", newline="") as file:
        writer = csv.writer(file)
        writer.writerows(matrix)
    print("Data loaded. Saved to", output_file)
    to_calendar_scheme(output_file, year, prefix)


def data_loader(workbook, offsets, prefix):
    print(f'--- Found *{len(offsets["CLASSES_SECTION"])}* years ---')
    sheet = workbook.active

    for k, v in offsets["CLASSES_SECTION"].items():
        print(f"Loading data for year {k}...")
        output_file = OUTPUT_FILE_CSV_LOAD + f"_{k}"
        matrix = []
        for start_col in range(
            offsets["CLASS_START_COL"], sheet.max_column, offsets["DATA_OFFSET"]
        ):
            value = sheet.cell(
                offsets["DATA"]["ROW"], column=start_col + offsets["DATA_CLASS_OFFSET"]
            ).value
            data = str(value).split(" ")[0]
            matrix.append([data])

            end_col = start_col + offsets["CLASS_END_COl"]
            for row in range(v[0] + 1, v[1] + 1):
                row_data = []
                for col in range(start_col, min(end_col + 1, sheet.max_column + 1)):
                    cell = sheet.cell(row=row, column=col).value
                    data = str(cell)
                    row_data.append(data)
                matrix.append(row_data)

        with open(output_file, "w", newline="") as file:
            writer = csv.writer(file)
            writer.writerows(matrix)
        print("Data loaded. Saved to", output_file)
        to_calendar_scheme(output_file, k, prefix)


def to_calendar_scheme(file_name, yearId, prefix):
    print(f"Creating CSV scheme from {file_name}...")
    with open(file_name, "r") as file:
        csv_reader = csv.reader(file)
        data = list(csv_reader)

    output_file = f"{prefix}_" + OUTPUT_FILE_CSV_CALENDAR + f"_{yearId}"
    with open(output_file, "w", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(
            [
                "Subject",
                "Start Date",
                "Start Time",
                "End Date",
                "End Time",
                "Description",
                "Location",
            ]
        )

    for i in range(len(data) - 2):
        for j in range(len(data[i])):
            if re.match(r"\d{4}-\d{2}-\d{2}", data[i][j]):
                for k in range(len(data[i + 1])):
                    if data[i + 1][k] != "None":
                        sala_index = k

                        if int(k + 1) % 2 == 0:
                            print(
                                "Found error in hour planning for day: ",
                                data[i][j],
                                "-- not writing to plan",
                            )
                            continue

                        subject = replace_polish_chars(data[i + 1][k])
                        teacher = replace_polish_chars(data[i + 2][k])
                        if teacher == "None":
                            teacher = "???"
                        sala = data[i + 6][sala_index + 1]
                        if sala == "None":
                            sala = "???"

                        with open(output_file, "a", newline="") as file:
                            writer = csv.writer(file)
                            writer.writerow(
                                [
                                    subject,
                                    data[i][j],
                                    retrieve_start(k + 1),
                                    data[i][j],
                                    retrieve_end(k + 1),
                                    teacher,
                                    sala,
                                ]
                            )
    os.remove(file_name)
    csv_to_ics(output_file, prefix, yearId)


wb = create_workbook(INPUT_FILE)

# get_calendar_schemes(wb)
# get_major_calendar_scheme(wb, "Informatyka", 4)
